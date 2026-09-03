"""
Reads a Genartml monthly timesheet workbook.

Layout expected per employee tab (matches the August 2026 file):
    B1  employee name          D1  default shift
    row 2  headers
    rows 3..33  daily data:
        A Date  B Day  C Type  D Shift  E Check In  F Check Out
        G Late  H Early Leave  I OT Hrs (device)  J Approved OT  K Description

Tabs named 'Rules' and 'Template' are skipped.
"""

import datetime as dt
import re
from decimal import Decimal

import openpyxl

SKIP_TABS = {"rules", "template", "payroll", "summary"}


def _hours(v):
    """Excel stores durations as time or timedelta. Return decimal hours."""
    if v is None:
        return None
    if isinstance(v, dt.timedelta):
        return Decimal(str(round(v.total_seconds() / 3600, 6)))
    if isinstance(v, dt.time):
        return Decimal(str(round(v.hour + v.minute / 60 + v.second / 3600, 6)))
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    return None


def _clock(v, on_date):
    """Return a full datetime. Bare times are anchored to on_date."""
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, dt.time):
        return dt.datetime.combine(on_date, v)
    return None


def _shift_window(shift_str, on_date, default="09:00 - 18:00"):
    s = str(shift_str or default)
    m = re.findall(r"(\d{1,2}):(\d{2})", s)
    if len(m) < 2:
        m = re.findall(r"(\d{1,2}):(\d{2})", default)
    start = dt.datetime.combine(on_date, dt.time(int(m[0][0]), int(m[0][1])))
    end = dt.datetime.combine(on_date, dt.time(int(m[1][0]), int(m[1][1])))
    if end <= start:
        end += dt.timedelta(days=1)
    return start, end


def read_workbook(path, rules, year=None, month=None):
    """Return {tab_name: {'name':.., 'shift':.., 'rows':[...]}}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for tab in wb.sheetnames:
        if tab.strip().lower() in SKIP_TABS:
            continue
        ws = wb[tab]
        name = ws["B1"].value
        if not name:
            continue
        default_shift = ws["D1"].value or "09:00 - 18:00"
        rows = []
        for r in range(3, ws.max_row + 1):
            d = ws.cell(r, 1).value
            raw_type = ws.cell(r, 3).value
            if d is None or raw_type is None:
                continue
            if isinstance(d, dt.datetime):
                d = d.date()
            if not isinstance(d, dt.date):
                continue
            if year and month and (d.year != year or d.month != month):
                continue

            t = rules.normalise_type(raw_type)
            shift = ws.cell(r, 4).value
            if isinstance(shift, str) and shift.startswith("="):
                shift = default_shift
            shift = shift or default_shift

            ci = _clock(ws.cell(r, 5).value, d)
            co = _clock(ws.cell(r, 6).value, d)
            # ---- overnight fix: a checkout before check-in rolled past midnight
            crossed = False
            if ci and co and co <= ci:
                co += dt.timedelta(days=1)
                crossed = True

            s_start, s_end = _shift_window(shift, d)
            if ci and ci > s_start:
                s_end = ci + (s_end - s_start)   # late arrival: shift end = CI + 9h

            device_ot = _hours(ws.cell(r, 9).value)
            if ci and co:
                calc = Decimal(str(round(max(
                    (co - s_end).total_seconds() / 3600, 0), 6)))
                device_ot = calc if device_ot is None or crossed else device_ot

            rows.append({
                "date": d,
                "type": t,
                "shift": str(shift),
                "check_in": ci,
                "check_out": co,
                "crossed_midnight": crossed,
                "device_ot_hours": device_ot,
                "approved_ot_hours": _hours(ws.cell(r, 10).value) or Decimal(0),
                "note": ws.cell(r, 11).value,
            })
        out[tab] = {"name": str(name).strip(), "shift": str(default_shift), "rows": rows}
    return out
