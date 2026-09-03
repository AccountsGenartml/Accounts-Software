#!/usr/bin/env python3
"""
Genartml payroll runner.

    python run_payroll.py --timesheet Aug.xlsx --month 2026-08

Outputs into ./out/<month>/ :
    payslips.html      one page per employee, print to PDF
    payroll_summary.xlsx
    review_flags.txt
"""

import argparse
import datetime as dt
import json
import sys
from decimal import Decimal
from pathlib import Path

from engine import CompanyCalendar, Rules, run_payroll, PayrollError
from timesheet import read_workbook
from payslip import render_all, MONTHS

HERE = Path(__file__).parent


def load_employees():
    p = HERE / "config" / "employees.json"
    if not p.exists():
        sys.exit(f"Missing {p}. Create it from employees.sample.json.")
    return json.loads(p.read_text())


def match_employee(tab_name, sheet_name, master):
    key = sheet_name.strip().lower()
    for e in master:
        names = [e["name"].lower()] + [a.lower() for a in e.get("aliases", [])]
        if key in names:
            return e
    for e in master:
        if e.get("tab") and e["tab"].strip().lower() == tab_name.strip().lower():
            return e
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--timesheet", required=True)
    ap.add_argument("--month", required=True, help="YYYY-MM")
    ap.add_argument("--out", default=str(HERE / "out"))
    ap.add_argument("--strict", action="store_true",
                    help="exit non-zero if any review flag is raised")
    a = ap.parse_args()

    year, month = (int(x) for x in a.month.split("-"))
    rules = Rules()
    cal = CompanyCalendar()
    bd = cal.breakdown(year, month)
    wd = bd["working_days"]

    print(f"\n  GENARTML PAYROLL — {MONTHS[month]} {year}")
    print("  " + "-" * 60)
    print(f"  Calendar days      {bd['calendar_days']}")
    print(f"  Weekly offs      - {bd['weekly_offs']}")
    print(f"  Public holidays  - {bd['public_holidays']}"
          + (f"   ({', '.join(h['name'] for h in bd['public_holiday_dates'])})"
             if bd['public_holiday_dates'] else ""))
    for h in bd["public_holidays_on_weekend"]:
        print(f"      note: {h['name']} ({h['date']}) falls on a weekend — no comp day, not subtracted")
    print(f"  WORKING DAYS     = {wd}   (same divisor for every employee)\n")

    sheets = read_workbook(a.timesheet, rules, year, month)
    master = load_employees()

    pairs, unmatched = [], []
    for tab, data in sheets.items():
        emp = match_employee(tab, data["name"], master)
        if not emp:
            unmatched.append(f"{tab} ({data['name']})")
            continue
        emp = dict(emp)
        emp["ctc_monthly"] = Decimal(str(emp["ctc_monthly"]))
        for k in ("date_of_joining", "exit_date"):
            if emp.get(k):
                emp["doj_display" if k == "date_of_joining" else "exit_display"] = \
                    dt.date.fromisoformat(emp[k]).strftime("%d-%m-%Y")
                emp[k] = dt.date.fromisoformat(emp[k])
        pairs.append((emp, data["rows"]))

    if unmatched:
        sys.exit("  No employee master entry for: " + ", ".join(unmatched)
                 + "\n  Add them to config/employees.json. Refusing to guess a CTC.")

    try:
        results = run_payroll(pairs, wd, rules, cal, year, month)
    except PayrollError as e:
        sys.exit(f"  HALTED: {e}")

    results.sort(key=lambda r: r["employee"].get("emp_id", ""))

    outdir = Path(a.out) / f"{year}-{month:02d}"
    outdir.mkdir(parents=True, exist_ok=True)

    w = max(len(r["employee"]["name"]) for r in results)
    print(f"  {'Employee'.ljust(w)}  {'Days':>6} {'OT h':>7} {'Gross':>12} {'PT':>6} {'NET':>12}")
    print("  " + "-" * (w + 48))
    tot = Decimal(0)
    for r in results:
        tot += r["net_pay"]
        print(f"  {r['employee']['name'].ljust(w)}  "
              f"{float(r['payable_days']):>6.2f} {float(r['ot_hours']):>7.2f} "
              f"{float(r['gross']):>12,.2f} {float(r['professional_tax']):>6.0f} "
              f"{float(r['net_pay']):>12,.2f}")
    print("  " + "-" * (w + 48))
    print(f"  {'TOTAL NET'.ljust(w)}  {'':>6} {'':>7} {'':>12} {'':>6} {float(tot):>12,.2f}\n")

    html = render_all(results, year, month, outdir / "payslips.html")
    write_summary(results, outdir / "payroll_summary.xlsx", year, month, bd)

    flags = [(r["employee"]["name"], f) for r in results for f in r["flags"]]
    lines = [f"GENARTML PAYROLL REVIEW — {MONTHS[month]} {year}", "=" * 58, ""]
    if not flags:
        lines.append("No review flags. Clean run.")
    for n, f in flags:
        lines.append(f"[{f['code']}] {n}"
                     + (f"  {f['date']}" if f.get("date") else "") + f"\n    {f['detail']}")
    (outdir / "review_flags.txt").write_text("\n".join(lines))

    if flags:
        print(f"  {len(flags)} item(s) flagged for review → review_flags.txt")
        for n, f in flags:
            print(f"    [{f['code']}] {n}" + (f" {f['date']}" if f.get("date") else ""))
        print()
    print(f"  Payslips  {html}")
    print(f"  Summary   {outdir / 'payroll_summary.xlsx'}\n")

    if a.strict and flags:
        sys.exit(1)


def write_summary(results, path, year, month, bd):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"
    A = "Arial"
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    INR = '"₹"#,##0.00;("₹"#,##0.00);-'

    ws["A1"] = f"GENARTML PVT. LTD. — PAYROLL {MONTHS[month].upper()} {year}"
    ws["A1"].font = Font(name=A, size=13, bold=True)
    ws["A3"] = (f"Working days {bd['working_days']} = {bd['calendar_days']} calendar "
                f"− {bd['weekly_offs']} weekly offs − {bd['public_holidays']} public holidays")
    ws["A3"].font = Font(name=A, size=9, italic=True, color="808080")

    hdr = ["Employee", "Emp ID", "CTC", "Basic", "Per Day", "Per Hour", "OT Rate",
           "Payable Days", "LOP Days", "OT Hrs", "Base Earned", "OT Payable",
           "Allowance", "Incentive", "Gross", "Prof. Tax", "Net Payable", "Flags"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(5, i, h)
        c.font = Font(name=A, size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="404040")
        c.border = box
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = 14 if i > 2 else 24
    ws.row_dimensions[5].height = 28

    r = 6
    for res in results:
        e = res["employee"]
        vals = [e["name"], e.get("emp_id", ""), float(e["ctc_monthly"]),
                float(res["base_salary"]), float(res["per_day"]), float(res["per_hour"]),
                float(res["ot_rate"]), float(res["payable_days"]), float(res["lop_days"]),
                float(res["ot_hours"]), float(res["base_earned"]), float(res["ot_payable"]),
                float(res["allowance_paid"]), float(res["incentive_paid"]),
                float(res["gross"]), float(res["professional_tax"]), float(res["net_pay"]),
                ", ".join(sorted({f["code"] for f in res["flags"]}))]
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v)
            c.font = Font(name=A, size=10, bold=(i == 17))
            c.border = box
            if i in (3, 4, 5, 6, 7, 11, 12, 13, 14, 15, 16, 17):
                c.number_format = INR
            if i in (8, 9, 10):
                c.number_format = "0.00"
        r += 1

    ws.cell(r, 1, "TOTAL").font = Font(name=A, size=10, bold=True)
    for i in (11, 12, 13, 14, 15, 16, 17):
        L = ws.cell(5, i).column_letter
        c = ws.cell(r, i, f"=SUM({L}6:{L}{r-1})")
        c.number_format = INR
        c.font = Font(name=A, size=10, bold=True)
        c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.border = box
    ws.freeze_panes = "C6"
    wb.save(path)


if __name__ == "__main__":
    main()
